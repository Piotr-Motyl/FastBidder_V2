import datetime
import uuid
from django.conf import settings
import os
from openpyxl import load_workbook
import magic  # pip install python-magic-bin  # alternatywna implementacja dla Windows !!! Inne wersje nie działają na Windowsie
from .models import UploadedFile


class FileManagementService:

    def upload_file(self, file, file_type):
        """
        Obsługuje wgrywanie pliku i zapisuje go w odpowiednim katalogu.

        Parametry:
            file: Obiekt pliku z request.FILES
            file_type: Typ pliku ('WF' lub 'REF')

        Zwraca:
            Słownik zawierający informacje o wgranym pliku
                {
                    "file_id": str,
                    "file_path": str,
                    "original_filename": str
                }

        Zgłasza:
            ValueError: Gdy plik nie przejdzie walidacji
            IOError: Gdy wystąpi błąd podczas zapisywania pliku
        """
        # Walidacja pliku
        is_valid = self.validate_uploaded_file(file)
        if not is_valid:
            raise ValueError(
                f"Nieprawidłowy plik: {file.name}. Sprawdź format i rozmiar pliku."
            )

        # Wybór odpowiedniego katalogu docelowego
        if file_type == "WF":
            upload_dir = settings.UPLOAD_WF_DIR
        elif file_type == "REF":
            upload_dir = settings.UPLOAD_REF_DIR
        else:
            raise ValueError(
                f"Nieprawidłowy typ pliku: {file_type}. Dozwolone: 'WF' lub 'REF'."
            )

        # Generowanie unikalnej nazwy pliku zachowując oryginalną nazwę
        filename = file.name
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(upload_dir, f"{timestamp}_{filename}")

        try:
            # Zapisanie pliku na dysku
            with open(file_path, "wb+") as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            # Zapisanie informacji o pliku w bazie danych
            uploaded_file = UploadedFile(
                file=file_path.replace(settings.MEDIA_ROOT, "").lstrip(
                    "/"
                ),  # Ścieżka względna do MEDIA_ROOT
                file_type=file_type,
                original_filename=filename,
            )
            uploaded_file.save()

            return {
                "file_id": str(uploaded_file.id),
                "file_path": file_path,
                "original_filename": filename,
            }

        except IOError as e:
            # Logowanie błędu dla łatwiejszego debugowania
            print(f"Błąd podczas zapisywania pliku: {str(e)}")
            raise IOError(f"Nie udało się zapisać pliku: {str(e)}")

    def validate_uploaded_file(self, file):
        """
        Waliduje wgrany plik (format, rozmiar, itp.)

        Parametry:
            file: Obiekt pliku z request.FILES

        Zwraca:
            True jeśli plik jest prawidłowy, False w przeciwnym razie
        """
        try:
            # Sprawdzanie rozmiaru pliku
            max_size_bytes = (
                settings.FILE_UPLOAD_MAX_SIZE_MB * 1024 * 1024
            )  # Konwersja MB na bajty
            if file.size > max_size_bytes:
                print(
                    f"Plik zbyt duży: {file.size} bajtów. Maksymalny rozmiar: {max_size_bytes} bajtów."
                )
                return False

            # Sprawdzanie typu MIME pliku
            # Odczyt początkowych bajtów pliku do identyfikacji typu MIME
            file_content = file.read(2048)  # Odczyt pierwszych 2048 bajtów
            file.seek(0)  # Przywrócenie wskaźnika na początek pliku

            mime_type = magic.from_buffer(file_content, mime=True)
            if mime_type not in settings.EXCEL_MIME_TYPES:
                print(
                    f"Nieprawidłowy typ pliku: {mime_type}. Dozwolone typy: {settings.EXCEL_MIME_TYPES}"
                )
                return False

            # Dodatkowo sprawdzamy czy plik jest otwieralny przez openpyxl
            try:
                # Zapisujemy plik tymczasowo aby openpyxl mógł go otworzyć
                temp_path = os.path.join(settings.MEDIA_ROOT, "temp_file.xlsx")
                with open(temp_path, "wb") as temp_file:
                    for chunk in file.chunks():
                        temp_file.write(chunk)
                file.seek(0)  # Przywrócenie wskaźnika na początek pliku

                # Próbujemy otworzyć plik za pomocą openpyxl
                workbook = load_workbook(temp_path, read_only=True)
                workbook.close()

                # Usuwamy plik tymczasowy
                os.remove(temp_path)
            except Exception as e:
                print(f"Nie można otworzyć pliku jako arkusz Excel: {str(e)}")
                # Usuwamy plik tymczasowy jeśli istnieje
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return False

            return True

        except Exception as e:
            print(f"Błąd podczas walidacji pliku: {str(e)}")
            return False

    def get_upload_directory(self, file_type):
        """
        Zwraca ścieżkę do katalogu, w którym ma być zapisany plik.

        Parametry:
            file_type (str): Typ pliku ('WF' lub 'REF')

        Zwraca:
            str: Ścieżka do katalogu

        Raises:
            ValueError: Jeśli podano nieprawidłowy typ pliku
        """

        # Główny katalog dla przechowywania plików
        media_root = settings.MEDIA_ROOT

        # Ustalenie podkatalogu na podstawie typu pliku
        if file_type == "WF":
            subdir = os.path.join("uploads", "wf")
        elif file_type == "REF":
            subdir = os.path.join("uploads", "ref")
        else:
            raise ValueError(
                f"Nieprawidłowy typ pliku: {file_type}. Dozwolone: 'WF' lub 'REF'"
            )

        # Pełna ścieżka do katalogu
        upload_dir = os.path.join(media_root, subdir)

        # Upewnienie się, że katalog istnieje
        os.makedirs(upload_dir, exist_ok=True)

        return upload_dir

    def generate_unique_filename(self, original_filename, file_type):
        """
        Generuje unikalną nazwę pliku w oparciu o oryginalną nazwę i typ.

        Parametry:
            original_filename: Oryginalna nazwa pliku
            file_type: Typ pliku ('WF' lub 'REF')

        Zwraca:
            Unikalna nazwa pliku
        """
        pass

    def validate_files(self, working_file_path: str, reference_file_path: str) -> bool:
        """
        Waliduje czy pliki istnieją i mają poprawny format Excel (.xlsx).

        Parametry:
            working_file_path (str): Ścieżka do pliku WF (Working File)
            reference_file_path (str): Ścieżka do pliku REF (Reference File)

        Zwraca:
            dict: Słownik zawierający wynik walidacji
                {
                    "valid": bool,  # True jeśli oba pliki są prawidłowe
                    "errors": list  # Lista błędów, pusta jeśli valid=True
                }

        Zgłasza:
            FileNotFoundError: Jeśli któryś z plików nie istnieje
            ValidationError: Jeśli któryś z plików ma nieprawidłowy format
        """
        validation_result = {"valid": True, "errors": []}

        # Sprawdzenie, czy obie ścieżki są niepuste
        if not working_file_path or not reference_file_path:
            validation_result["valid"] = False
            validation_result["errors"].append("Ścieżki do plików nie mogą być puste")
            return validation_result

        # Lista plików do walidacji
        files_to_validate = [
            {"path": working_file_path, "type": "Working File (WF)"},
            {"path": reference_file_path, "type": "Reference File (REF)"},
        ]

        # Walidacja każdego pliku
        for file_info in files_to_validate:
            file_path = file_info["path"]
            file_type = file_info["type"]

            # Sprawdzenie czy plik istnieje
            if not self._file_exists(file_path):
                validation_result["valid"] = False
                validation_result["errors"].append(
                    f"Plik {file_type} nie istnieje: {file_path}"
                )
                continue

            # Sprawdzenie czy plik ma właściwy format
            if not self._is_valid_excel_file(file_path):
                validation_result["valid"] = False
                validation_result["errors"].append(
                    f"Plik {file_type} nie jest prawidłowym plikiem Excel: {file_path}"
                )
                continue

            # Sprawdzenie czy plik można otworzyć
            if not self._can_open_excel_file(file_path):
                validation_result["valid"] = False
                validation_result["errors"].append(
                    f"Nie można otworzyć pliku {file_type}: {file_path}"
                )
                continue

        return validation_result

    def _file_exists(self, file_path):
        """
        Sprawdza czy plik istnieje w systemie plików.

        Parametry:
            file_path (str): Ścieżka do pliku

        Zwraca:
            bool: True jeśli plik istnieje, False w przeciwnym razie
        """
        return os.path.isfile(file_path)

    def _is_valid_excel_file(self, file_path):
        """
        Sprawdza czy plik jest prawidłowym plikiem Excel (.xlsx).

        Parametry:
            file_path (str): Ścieżka do pliku

        Zwraca:
            bool: True jeśli plik jest prawidłowym plikiem Excel, False w przeciwnym razie
        """
        # Sprawdzenie rozszerzenia pliku
        if not file_path.lower().endswith(".xlsx"):
            return False

        # Sprawdzenie typu MIME
        mime = magic.Magic(mime=True)
        file_type = mime.from_file(file_path)
        valid_excel_types = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/zip",  # Excel .xlsx pliki są zaszyfrowanymi archiwami ZIP
        ]

        return file_type in valid_excel_types

    def _can_open_excel_file(self, file_path):
        """
        Sprawdza czy plik Excel można otworzyć (nie jest uszkodzony ani zablokowany).

        Parametry:
            file_path (str): Ścieżka do pliku

        Zwraca:
            bool: True jeśli plik można otworzyć, False w przeciwnym razie
        """
        try:
            # Próba otwarcia pliku za pomocą openpyxl
            workbook = load_workbook(filename=file_path, read_only=True)
            workbook.close()
            return True
        except Exception:
            return False

    def get_file_paths(self, working_file_identifier, reference_file_identifier):
        """
        Zwraca pełne ścieżki do plików WF i REF na podstawie ich identyfikatorów.

        Identyfikator może być:
        - UUID pliku w bazie danych (jako string lub obiekt UUID)
        - Ścieżką do pliku (relatywną lub absolutną)

        Parametry:
            working_file_identifier (str/UUID): Identyfikator pliku WF
            reference_file_identifier (str/UUID): Identyfikator pliku REF

        Zwraca:
            dict: Słownik zawierający ścieżki do plików
                {
                    "working_file_path": str,  # Pełna ścieżka do pliku WF
                    "reference_file_path": str  # Pełna ścieżka do pliku REF
                }

        Zgłasza:
            FileNotFoundError: Jeśli któryś z plików nie istnieje
            ValueError: Jeśli identyfikator jest nieprawidłowy
        """
        result = {}

        # Pobierz ścieżki do plików
        result["working_file_path"] = self._resolve_file_path(
            working_file_identifier, "WF"
        )
        result["reference_file_path"] = self._resolve_file_path(
            reference_file_identifier, "REF"
        )

        # Weryfikacja istnienia plików
        for key, path in result.items():
            if not os.path.isfile(path):
                file_type = (
                    "Working File" if key == "working_file_path" else "Reference File"
                )
                raise FileNotFoundError(f"{file_type} nie istnieje pod ścieżką: {path}")

        return result

    def _resolve_file_path(self, file_identifier, file_type):
        """
        Metoda pomocnicza rozwiązująca ścieżkę do pliku na podstawie identyfikatora.

        Parametry:
            file_identifier (str/UUID): Identyfikator pliku
            file_type (str): Typ pliku ('WF' lub 'REF')

        Zwraca:
            str: Pełna ścieżka do pliku

        Zgłasza:
            ValueError: Jeśli identyfikator jest nieprawidłowy lub nie można znaleźć pliku
        """
        # Sprawdź czy identyfikator jest pusty
        if not file_identifier:
            raise ValueError(f"Identyfikator pliku {file_type} nie może być pusty")

        # Sprawdź czy to bezpośrednia ścieżka do pliku
        if isinstance(file_identifier, str) and os.path.isfile(file_identifier):
            return os.path.abspath(file_identifier)

        # Spróbuj zinterpretować jako UUID i znaleźć plik w bazie danych
        try:
            if isinstance(file_identifier, str):
                file_id = uuid.UUID(file_identifier)
            else:
                file_id = file_identifier

            file_obj = UploadedFile.objects.get(id=file_id, file_type=file_type)
            return file_obj.get_file_path()
        except (ValueError, UploadedFile.DoesNotExist):
            pass

        # Sprawdź czy to nazwa pliku, i znajdź plik w odpowiednim katalogu
        if isinstance(file_identifier, str):
            base_dir = self.get_upload_directory(file_type)
            potential_path = os.path.join(base_dir, file_identifier)
            if os.path.isfile(potential_path):
                return potential_path

        # Jeśli dotarliśmy tutaj, nie udało się znaleźć pliku
        raise ValueError(
            f"Nie można znaleźć pliku {file_type} o identyfikatorze: {file_identifier}"
        )

    def get_result_file_path(self, working_file_path: str) -> str:
        # TODO: zmienić, aby resultat lądował w media/results zamiast tam, skąd pochodzi plik WF.
        # Zaharkodowane przeniesienie do media/results na końcu metody
        """
        Generuje ścieżkę do pliku wynikowego na podstawie ścieżki do pliku roboczego.

        Proces obejmuje:
        1. Walidację ścieżki do pliku roboczego
        2. Utworzenie nowej ścieżki dla pliku wynikowego, w tym samym katalogu co plik roboczy,
           ale z dodatkowym oznaczeniem (suffixem "_result_YYYYMMDD-HHMM")
        3. Zapewnienie, że katalog docelowy istnieje i jest dostępny do zapisu

        Args:
            working_file_path (str): Ścieżka do oryginalnego pliku roboczego (WF)

        Returns:
            str: Ścieżka do pliku wynikowego, gdzie zostanie zapisany zmodyfikowany plik WF

        Raises:
            ValueError: Gdy ścieżka do pliku roboczego jest nieprawidłowa
            PermissionError: Gdy brak uprawnień do zapisu w katalogu docelowym
            IOError: Gdy wystąpi inny problem z systemem plików
        """
        # 1. Walidacja ścieżki do pliku roboczego
        if not working_file_path:
            raise ValueError("Ścieżka do pliku roboczego nie może być pusta")

        if not os.path.isfile(working_file_path):
            raise ValueError(f"Plik roboczy nie istnieje: {working_file_path}")

        if not working_file_path.lower().endswith(".xlsx"):
            raise ValueError(
                f"Plik roboczy musi być w formacie Excel (.xlsx): {working_file_path}"
            )

        # 2. Utworzenie nowej ścieżki dla pliku wynikowego
        # Pobierz katalog i nazwę pliku
        file_dir = os.path.dirname(working_file_path)
        file_name = os.path.basename(working_file_path)

        # Oddziel nazwę pliku od rozszerzenia
        file_base, file_ext = os.path.splitext(file_name)

        # Generuj znacznik czasowy w formacie YYYYMMDD-HHMM
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")

        # Utwórz nazwę pliku wynikowego z suffixem
        result_file_name = f"{file_base}_result_{timestamp}{file_ext}"

        # Zbuduj pełną ścieżkę
        result_file_path = os.path.join(file_dir, result_file_name)

        # 3. Sprawdź, czy katalog docelowy istnieje i jest dostępny do zapisu
        try:
            # Sprawdź, czy katalog istnieje
            if not os.path.isdir(file_dir):
                raise IOError(f"Katalog docelowy nie istnieje: {file_dir}")

            # Sprawdź uprawnienia do zapisu poprzez próbę utworzenia pustego pliku testowego
            test_file_path = os.path.join(
                file_dir, f".test_write_permission_{timestamp}"
            )
            try:
                with open(test_file_path, "w") as f:
                    pass
                os.remove(test_file_path)  # Usuń plik testowy po udanym teście
            except (IOError, PermissionError) as e:
                raise PermissionError(
                    f"Brak uprawnień do zapisu w katalogu docelowym: {file_dir}. Błąd: {str(e)}"
                )

        except IOError as e:
            raise IOError(f"Problem z dostępem do systemu plików: {str(e)}")

        # TESTOWA MODYFIKACJA: Nadpisanie ścieżki pliku wynikowego na media/results
        # Możesz zakomentować poniższe dwie linie, aby wrócić do oryginalnej funkcjonalności
        results_dir = os.path.join("media", "results")
        result_file_path = os.path.join(results_dir, result_file_name)

        # Upewnij się, że katalog docelowy istnieje
        os.makedirs(results_dir, exist_ok=True)

        return result_file_path
